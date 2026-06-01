import csv
from datetime import timedelta
from io import StringIO

from django.contrib import messages as _messages
from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required


class _DashboardMessages:
    """Обёртка над django.contrib.messages, которая помечает все сообщения
    тегом 'dashboard' — чтобы они отображались только внутри админ-панели
    и не утекали на публичные страницы (если пользователь сразу перейдёт туда)."""

    @staticmethod
    def _add(level_func, request, msg, extra_tags=""):
        tag = "dashboard"
        if extra_tags:
            tag = f"{tag} {extra_tags}"
        level_func(request, msg, extra_tags=tag)

    @classmethod
    def success(cls, request, msg, extra_tags=""):
        cls._add(_messages.success, request, msg, extra_tags)

    @classmethod
    def info(cls, request, msg, extra_tags=""):
        cls._add(_messages.info, request, msg, extra_tags)

    @classmethod
    def error(cls, request, msg, extra_tags=""):
        cls._add(_messages.error, request, msg, extra_tags)

    @classmethod
    def warning(cls, request, msg, extra_tags=""):
        cls._add(_messages.warning, request, msg, extra_tags)


messages = _DashboardMessages
from django.core.paginator import Paginator
from django.db import models
from django.db.models import Count, Sum
from django.http import HttpResponseRedirect
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views import View

from apps.accounts.models import User
from apps.catalog.models import Category, DigitalItem, Product, ProductMedia, Review
from apps.core.models import SecurityEvent, SiteSetting, SupportMessage, SupportTicket
from apps.orders.models import Order

from .forms import (
    CategoryForm,
    DigitalItemForm,
    OrderForm,
    ProductForm,
    ProductMediaForm,
    SiteSettingsForm,
    UserForm,
)


def _log_admin_action(request, description: str) -> None:
    """Record admin action in SecurityEvent."""
    SecurityEvent.objects.create(
        event_type="admin_action",
        description=description,
        user=request.user if request.user.is_authenticated else None,
        ip_address=_get_client_ip(request),
        user_agent=(request.META.get("HTTP_USER_AGENT") or "")[:512],
    )


def _get_client_ip(request) -> str | None:
    xff = request.META.get("HTTP_X_FORWARDED_FOR")
    if xff:
        return xff.split(",")[0].strip() or None
    return request.META.get("REMOTE_ADDR") or None


@method_decorator(staff_member_required, name="dispatch")
class DashboardView(View):
    template_name = "dashboard/index.html"

    def get(self, request):
        import json
        from django.db.models.functions import TruncDate

        now = timezone.now()
        week_ago = now - timedelta(days=7)
        month_ago = now - timedelta(days=30)

        orders_week = Order.objects.filter(created_at__gte=week_ago)
        revenue_week = orders_week.filter(status=Order.Status.PAID).aggregate(
            total=Sum("amount")
        )["total"] or 0
        orders_count_week = orders_week.count()
        total_orders = Order.objects.count()
        total_revenue = Order.objects.filter(status=Order.Status.PAID).aggregate(
            total=Sum("amount")
        )["total"] or 0
        total_users = User.objects.count()
        total_products = Product.objects.filter(status=Product.Status.ACTIVE).count()
        top_products = (
            Product.objects.filter(order_items__order__status=Order.Status.PAID)
            .annotate(cnt=Count("order_items"))
            .order_by("-cnt")[:5]
        )
        latest_orders = Order.objects.select_related("user").order_by("-created_at")[:10]

        open_tickets_count = SupportTicket.objects.filter(
            status=SupportTicket.Status.OPEN
        ).count()
        recent_tickets = (
            SupportTicket.objects.select_related("user")
            .order_by("-updated_at")[:5]
        )
        pending_reviews_count = Review.objects.filter(
            status=Review.Status.PENDING
        ).count()

        # ===== Данные для графиков (Chart.js) =====

        # 1) Выручка и количество оплаченных заказов по дням за 30 дней
        by_day_raw = (
            Order.objects.filter(
                created_at__gte=month_ago,
                status=Order.Status.PAID,
            )
            .annotate(day=TruncDate("created_at"))
            .values("day")
            .annotate(revenue=Sum("amount"), cnt=Count("id"))
            .order_by("day")
        )
        day_map = {r["day"]: r for r in by_day_raw}
        labels_30d = []
        revenue_30d = []
        orders_30d = []
        for i in range(29, -1, -1):
            d = (now - timedelta(days=i)).date()
            labels_30d.append(d.strftime("%d.%m"))
            rec = day_map.get(d)
            revenue_30d.append(float(rec["revenue"]) if rec else 0.0)
            orders_30d.append(int(rec["cnt"]) if rec else 0)

        # 2) Распределение по статусам
        status_counts_raw = (
            Order.objects.values("status").annotate(cnt=Count("id"))
        )
        status_labels_map = dict(Order.Status.choices)
        status_labels = []
        status_values = []
        for row in status_counts_raw:
            status_labels.append(str(status_labels_map.get(row["status"], row["status"])))
            status_values.append(int(row["cnt"]))

        # 3) Топ-5 товаров по выручке
        top_revenue_raw = (
            Product.objects.filter(order_items__order__status=Order.Status.PAID)
            .annotate(
                revenue=Sum(
                    models.F("order_items__price") * models.F("order_items__quantity"),
                    output_field=models.DecimalField(max_digits=14, decimal_places=2),
                ),
            )
            .order_by("-revenue")[:5]
            .values("name", "revenue")
        )
        top_rev_labels = [r["name"] for r in top_revenue_raw]
        top_rev_values = [float(r["revenue"] or 0) for r in top_revenue_raw]

        charts = {
            "labels_30d": json.dumps(labels_30d),
            "revenue_30d": json.dumps(revenue_30d),
            "orders_30d": json.dumps(orders_30d),
            "status_labels": json.dumps(status_labels, ensure_ascii=False),
            "status_values": json.dumps(status_values),
            "top_rev_labels": json.dumps(top_rev_labels, ensure_ascii=False),
            "top_rev_values": json.dumps(top_rev_values),
        }

        # 4) Конверсия в оплату за 30 дней
        month_orders = Order.objects.filter(created_at__gte=month_ago)
        month_total = month_orders.count()
        month_paid = month_orders.filter(status=Order.Status.PAID).count()
        conversion = round((month_paid / month_total * 100), 1) if month_total else 0

        return render(
            request,
            self.template_name,
            {
                "orders_count_week": orders_count_week,
                "revenue_week": revenue_week,
                "total_orders": total_orders,
                "total_revenue": total_revenue,
                "total_users": total_users,
                "total_products": total_products,
                "top_products": top_products,
                "latest_orders": latest_orders,
                "open_tickets_count": open_tickets_count,
                "recent_tickets": recent_tickets,
                "pending_reviews_count": pending_reviews_count,
                "charts": charts,
                "month_conversion": conversion,
                "month_orders_total": month_total,
                "month_orders_paid": month_paid,
            },
        )


@method_decorator(staff_member_required, name="dispatch")
class ProductsView(View):
    template_name = "dashboard/products.html"

    def get(self, request):
        qs = Product.objects.select_related("category").order_by("-created_at")
        q = request.GET.get("q", "").strip()
        status_filter = request.GET.get("status")
        if q:
            qs = qs.filter(
                models.Q(name__icontains=q)
                | models.Q(short_description__icontains=q)
            )
        if status_filter:
            qs = qs.filter(status=status_filter)
        paginator = Paginator(qs, 20)
        page = request.GET.get("page", 1)
        products = paginator.get_page(page)
        return render(
            request,
            self.template_name,
            {
                "products": products,
                "paginator": paginator,
                "q": q,
                "status_filter": status_filter,
            },
        )


@method_decorator(staff_member_required, name="dispatch")
class ProductEditView(View):
    template_name = "dashboard/product_edit.html"

    def get(self, request, pk=None):
        product = (
            Product.objects.filter(pk=pk).prefetch_related(
                "digital_items", "media"
            ).first()
            if pk
            else None
        )
        form = ProductForm(instance=product)
        digital_form = DigitalItemForm() if product else None
        media_form = ProductMediaForm() if product else None
        return render(
            request,
            self.template_name,
            {
                "form": form,
                "product": product,
                "digital_form": digital_form,
                "media_form": media_form,
            },
        )

    def post(self, request, pk=None):
        product = Product.objects.filter(pk=pk).prefetch_related(
            "digital_items", "media"
        ).first()
        if request.POST.get("delete_product") and product:
            product.delete()
            messages.success(request, "Товар удалён.")
            return redirect("dashboard:products")
        if request.POST.get("add_digital") and product:
            digital_form = DigitalItemForm(request.POST)
            if digital_form.is_valid():
                obj = digital_form.save(commit=False)
                obj.product = product
                obj.save()
                messages.success(request, "Цифровой элемент добавлен.")
                return redirect("dashboard:product_edit", pk=product.pk)
        elif request.POST.get("add_media") and product:
            media_form = ProductMediaForm(request.POST)
            if media_form.is_valid():
                obj = media_form.save(commit=False)
                obj.product = product
                obj.save()
                messages.success(request, "Медиа добавлено.")
                return redirect("dashboard:product_edit", pk=product.pk)
        else:
            form = ProductForm(request.POST, instance=product)
            if form.is_valid():
                obj = form.save()
                if not product:
                    messages.success(request, "Товар создан. Добавьте цифровые элементы и медиа.")
                    return redirect("dashboard:product_edit", pk=obj.pk)
                messages.success(request, "Товар сохранён.")
                return redirect("dashboard:products")
            digital_form = DigitalItemForm() if product else None
            media_form = ProductMediaForm() if product else None
            return render(
                request,
                self.template_name,
                {
                    "form": form,
                    "product": product,
                    "digital_form": digital_form,
                    "media_form": media_form,
                },
            )
        digital_form = (
            DigitalItemForm(request.POST)
            if request.POST.get("add_digital") and product
            else (DigitalItemForm() if product else None)
        )
        media_form = (
            ProductMediaForm(request.POST)
            if request.POST.get("add_media") and product
            else (ProductMediaForm() if product else None)
        )
        return render(
            request,
            self.template_name,
            {
                "form": ProductForm(instance=product),
                "product": product,
                "digital_form": digital_form,
                "media_form": media_form,
            },
        )


@method_decorator(staff_member_required, name="dispatch")
class CategoriesView(View):
    template_name = "dashboard/categories.html"

    def get(self, request):
        categories = Category.objects.order_by("sort_order", "name")
        return render(request, self.template_name, {"categories": categories})


@method_decorator(staff_member_required, name="dispatch")
class CategoryEditView(View):
    template_name = "dashboard/category_edit.html"

    def get(self, request, pk=None):
        category = Category.objects.filter(pk=pk).first()
        form = CategoryForm(instance=category)
        return render(request, self.template_name, {"form": form, "category": category})

    def post(self, request, pk=None):
        category = Category.objects.filter(pk=pk).first()
        if request.POST.get("delete") and category:
            name, cid = category.name, category.pk
            category.delete()
            _log_admin_action(request, f"Удаление категории: {name} (id={cid})")
            messages.success(request, "Категория удалена.")
            return redirect("dashboard:categories")
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, "Категория сохранена.")
            return redirect("dashboard:categories")
        return render(request, self.template_name, {"form": form, "category": category})


@method_decorator(staff_member_required, name="dispatch")
class UsersView(View):
    template_name = "dashboard/users.html"

    def get(self, request):
        qs = User.objects.order_by("-date_joined")
        q = request.GET.get("q", "").strip()
        if q:
            qs = qs.filter(
                models.Q(email__icontains=q)
                | models.Q(first_name__icontains=q)
                | models.Q(last_name__icontains=q)
            )
        paginator = Paginator(qs, 25)
        page = request.GET.get("page", 1)
        users = paginator.get_page(page)
        return render(
            request,
            self.template_name,
            {"users": users, "paginator": paginator, "q": q},
        )


@method_decorator(staff_member_required, name="dispatch")
class UserEditView(View):
    template_name = "dashboard/user_edit.html"

    def get(self, request, pk):
        user = get_object_or_404(User, pk=pk)
        form = UserForm(instance=user)
        return render(request, self.template_name, {"form": form, "user_obj": user})

    def post(self, request, pk):
        user = get_object_or_404(User, pk=pk)
        form = UserForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, "Пользователь сохранён.")
            return redirect("dashboard:users")
        return render(request, self.template_name, {"form": form, "user_obj": user})


@method_decorator(staff_member_required, name="dispatch")
class OrdersView(View):
    template_name = "dashboard/orders.html"

    def _filtered_qs(self, request):
        status = request.GET.get("status")
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        q = (request.GET.get("q") or "").strip()
        qs = Order.objects.select_related("user").order_by("-created_at")
        if status:
            qs = qs.filter(status=status)
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        if q:
            qs = qs.filter(
                models.Q(id__icontains=q)
                | models.Q(user__email__icontains=q)
            )
        return qs

    def get(self, request):
        qs = self._filtered_qs(request)
        paginator = Paginator(qs, 25)
        page = request.GET.get("page", 1)
        orders = paginator.get_page(page)
        return render(
            request,
            self.template_name,
            {
                "orders": orders,
                "paginator": paginator,
                "status_choices": Order.Status.choices,
                "status_filter": request.GET.get("status"),
                "date_from": request.GET.get("date_from"),
                "date_to": request.GET.get("date_to"),
                "q": request.GET.get("q", ""),
            },
        )

    def post(self, request):
        """Массовые операции: отметить paid / cancelled / fulfilled по ID."""
        action = request.POST.get("bulk_action")
        ids = [int(x) for x in request.POST.getlist("order_ids") if x.isdigit()]
        if not ids or action not in ("paid", "cancelled", "fulfilled"):
            messages.error(request, "Не выбраны заказы или неизвестное действие.")
            return redirect("dashboard:orders")
        mapping = {
            "paid": Order.Status.PAID,
            "cancelled": Order.Status.CANCELLED,
            "fulfilled": Order.Status.FULFILLED,
        }
        new_status = mapping[action]
        updated = 0
        for o in Order.objects.filter(pk__in=ids):
            if o.status != new_status:
                o.status = new_status
                o.save(update_fields=["status"])
                updated += 1
        _log_admin_action(
            request,
            f"Массовая смена статуса ({len(ids)} заказов) → {new_status}; обновлено {updated}",
        )
        messages.success(request, f"Обновлено заказов: {updated}.")
        return redirect("dashboard:orders")


@method_decorator(staff_member_required, name="dispatch")
class OrdersExportView(View):
    """Экспорт заказов в CSV/XLSX/PDF с теми же фильтрами, что и список."""

    def _filtered_qs(self, request):
        status = request.GET.get("status")
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        qs = Order.objects.select_related("user").order_by("-created_at")
        if status:
            qs = qs.filter(status=status)
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        return qs

    def get(self, request):
        fmt = (request.GET.get("format") or "csv").lower()
        qs = self._filtered_qs(request)
        if fmt == "xlsx":
            return self._export_xlsx(qs)
        if fmt == "pdf":
            return self._export_pdf(qs)
        return self._export_csv(qs)

    def _rows(self, qs):
        for o in qs:
            yield [
                o.id,
                o.user.email if o.user else "",
                o.get_status_display(),
                float(o.amount or 0),
                o.currency,
                o.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            ]

    def _export_csv(self, qs):
        buf = StringIO()
        writer = csv.writer(buf)
        writer.writerow(["ID", "Email", "Статус", "Сумма", "Валюта", "Создан"])
        for row in self._rows(qs):
            writer.writerow(row)
        response = HttpResponse("\ufeff" + buf.getvalue(), content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="orders_export.csv"'
        return response

    def _export_xlsx(self, qs):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            messages.error(self.request, "Библиотека openpyxl не установлена.")
            return redirect("dashboard:orders")
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"
        headers = ["ID", "Email", "Статус", "Сумма", "Валюта", "Создан"]
        ws.append(headers)
        header_fill = PatternFill("solid", fgColor="4F46E5")
        bold_white = Font(bold=True, color="FFFFFF")
        for col_idx, _ in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = header_fill
            c.font = bold_white
            c.alignment = Alignment(horizontal="center")
        for row in self._rows(qs):
            ws.append(row)
        widths = [8, 28, 18, 14, 8, 22]
        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w
        from io import BytesIO
        bio = BytesIO()
        wb.save(bio)
        response = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="orders_export.xlsx"'
        return response

    def _export_pdf(self, qs):
        try:
            from io import BytesIO
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        except ImportError:
            messages.error(self.request, "Библиотека reportlab не установлена.")
            return redirect("dashboard:orders")

        # Регистрируем кириллический шрифт (DejaVuSans из пакета reportlab).
        font_name = "Helvetica"
        try:
            import os
            from reportlab import __path__ as rlpath
            candidate = os.path.join(rlpath[0], "fonts", "DejaVuSans.ttf")
            if os.path.exists(candidate):
                pdfmetrics.registerFont(TTFont("DejaVuSans", candidate))
                font_name = "DejaVuSans"
        except Exception:
            pass

        bio = BytesIO()
        doc = SimpleDocTemplate(bio, pagesize=landscape(A4), title="Orders export")
        styles = getSampleStyleSheet()
        title_style = styles["Title"]
        title_style.fontName = font_name
        small_style = styles["Normal"]
        small_style.fontName = font_name

        story = [
            Paragraph("Экспорт заказов ShoShop", title_style),
            Paragraph(f"Сгенерировано: {timezone.now().strftime('%Y-%m-%d %H:%M')}", small_style),
            Spacer(1, 12),
        ]
        data = [["ID", "Email", "Статус", "Сумма", "Валюта", "Создан"]]
        for row in self._rows(qs):
            data.append([str(c) for c in row])
        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, -1), font_name),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                ]
            )
        )
        story.append(table)
        doc.build(story)
        response = HttpResponse(bio.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="orders_export.pdf"'
        return response


@method_decorator(staff_member_required, name="dispatch")
class OrderEditView(View):
    template_name = "dashboard/order_edit.html"

    def get(self, request, pk):
        order = get_object_or_404(Order, pk=pk)
        form = OrderForm(instance=order)
        return render(request, self.template_name, {"form": form, "order": order})

    def post(self, request, pk):
        order = get_object_or_404(Order, pk=pk)
        old_status = order.status
        form = OrderForm(request.POST, instance=order)
        if form.is_valid():
            form.save()
            if order.status != old_status:
                status_labels = dict(Order.Status.choices)
                _log_admin_action(
                    request,
                    f"Смена статуса заказа #{order.id}: "
                    f"{status_labels.get(old_status, old_status)} → {order.get_status_display()}",
                )
            messages.success(request, "Статус заказа обновлён.")
            return redirect("dashboard:order_detail", pk=pk)
        return render(request, self.template_name, {"form": form, "order": order})


@method_decorator(staff_member_required, name="dispatch")
class OrderDetailView(View):
    template_name = "dashboard/order_detail.html"

    def get(self, request, pk):
        order = get_object_or_404(
            Order.objects.select_related("user").prefetch_related("items__product"),
            pk=pk,
        )
        try:
            from apps.orders.models import OrderEvent
            events = list(
                OrderEvent.objects.filter(order=order)
                .select_related("actor")
                .order_by("-created_at")
            )
        except Exception:
            events = []
        return render(
            request, self.template_name,
            {"order": order, "events": events},
        )

    def post(self, request, pk):
        """Быстрое действие: добавить внутренний комментарий к заказу."""
        order = get_object_or_404(Order, pk=pk)
        note = (request.POST.get("note") or "").strip()
        if note:
            try:
                from apps.orders.models import OrderEvent
                OrderEvent.objects.create(
                    order=order,
                    event_type=OrderEvent.EventType.NOTE,
                    description=note[:500],
                    actor=request.user,
                )
                messages.success(request, "Комментарий добавлен.")
            except Exception:
                messages.error(request, "Не удалось добавить комментарий.")
        return redirect("dashboard:order_detail", pk=pk)


@method_decorator(staff_member_required, name="dispatch")
class ProductDeleteView(View):
    def post(self, request, pk):
        product = get_object_or_404(Product, pk=pk)
        name, pid = product.name, product.pk
        product.delete()
        _log_admin_action(request, f"Удаление товара: {name} (id={pid})")
        messages.success(request, "Товар удалён.")
        return redirect("dashboard:products")


@method_decorator(staff_member_required, name="dispatch")
class ProductDeleteDigitalView(View):
    def post(self, request, product_pk, item_pk):
        DigitalItem.objects.filter(product_id=product_pk, pk=item_pk).delete()
        messages.info(request, "Цифровой элемент удалён.")
        return redirect("dashboard:product_edit", pk=product_pk)


@method_decorator(staff_member_required, name="dispatch")
class ProductDeleteMediaView(View):
    def post(self, request, product_pk, media_pk):
        ProductMedia.objects.filter(product_id=product_pk, pk=media_pk).delete()
        messages.info(request, "Медиа удалено.")
        return redirect("dashboard:product_edit", pk=product_pk)


@method_decorator(staff_member_required, name="dispatch")
class SettingsView(View):
    template_name = "dashboard/settings.html"

    def get(self, request):
        form = SiteSettingsForm()
        return render(request, self.template_name, {"form": form})

    def post(self, request):
        form = SiteSettingsForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Настройки сохранены.")
            return redirect("dashboard:settings")
        return render(request, self.template_name, {"form": form})


@method_decorator(staff_member_required, name="dispatch")
class SecurityLogView(View):
    template_name = "dashboard/security_log.html"

    def get(self, request):
        qs = SecurityEvent.objects.select_related("user").order_by("-created_at")
        event_type = request.GET.get("event_type")
        if event_type:
            qs = qs.filter(event_type=event_type)
        paginator = Paginator(qs, 50)
        page = request.GET.get("page", 1)
        events = paginator.get_page(page)
        return render(
            request,
            self.template_name,
            {
                "events": events,
                "paginator": paginator,
                "event_type_filter": event_type,
                "event_types": SecurityEvent.EVENT_TYPES,
            },
        )


@method_decorator(staff_member_required, name="dispatch")
class ReviewsView(View):
    template_name = "dashboard/reviews.html"

    def get(self, request):
        qs = Review.objects.select_related("product", "user").order_by("-created_at")
        status_filter = request.GET.get("status")
        if status_filter:
            qs = qs.filter(status=status_filter)
        paginator = Paginator(qs, 30)
        page = request.GET.get("page", 1)
        reviews = paginator.get_page(page)
        return render(
            request,
            self.template_name,
            {
                "reviews": reviews,
                "paginator": paginator,
                "status_filter": status_filter,
                "status_choices": Review.Status.choices,
            },
        )


@method_decorator(staff_member_required, name="dispatch")
class ReviewModerateView(View):
    """POST: review_id, action=approve|hide."""

    def post(self, request):
        review_id = request.POST.get("review_id")
        action = request.POST.get("action")
        if not review_id or action not in ("approve", "hide"):
            messages.error(request, "Неверный запрос.")
            return redirect("dashboard:reviews")
        review = get_object_or_404(Review, pk=review_id)
        if action == "approve":
            review.status = Review.Status.PUBLISHED
            review.save(update_fields=["status"])
            messages.success(request, "Отзыв одобрен.")
        else:
            review.status = Review.Status.HIDDEN
            review.save(update_fields=["status"])
            messages.success(request, "Отзыв скрыт.")
        base = reverse("dashboard:reviews")
        qs = request.GET.urlencode()
        if qs:
            base = f"{base}?{qs}"
        return HttpResponseRedirect(base)


@method_decorator(staff_member_required, name="dispatch")
class SupportTicketsView(View):
    template_name = "dashboard/support_list.html"

    def get(self, request):
        qs = SupportTicket.objects.select_related("user").prefetch_related("messages").order_by("-updated_at")
        status_filter = request.GET.get("status")
        if status_filter and status_filter in ("open", "completed"):
            qs = qs.filter(status=status_filter)
        paginator = Paginator(qs, 20)
        page = request.GET.get("page", 1)
        tickets = paginator.get_page(page)
        return render(
            request,
            self.template_name,
            {
                "tickets": tickets,
                "paginator": paginator,
                "status_filter": status_filter,
                "status_choices": SupportTicket.Status.choices,
            },
        )


@method_decorator(staff_member_required, name="dispatch")
class SupportTicketDetailView(View):
    template_name = "dashboard/support_detail.html"

    def get(self, request, pk):
        ticket = get_object_or_404(SupportTicket.objects.select_related("user"), pk=pk)
        msgs = ticket.messages.select_related("author").order_by("created_at")
        return render(request, self.template_name, {"ticket": ticket, "messages": msgs})

    def post(self, request, pk):
        ticket = get_object_or_404(SupportTicket, pk=pk)
        action = request.POST.get("action")

        if action == "complete":
            if ticket.status != SupportTicket.Status.COMPLETED:
                ticket.status = SupportTicket.Status.COMPLETED
                ticket.save(update_fields=["status"])
                _log_admin_action(request, f"Обращение в поддержку #{ticket.pk} помечено завершённым.")
                messages.success(request, "Обращение помечено как завершённое.")
            return redirect("dashboard:support_detail", pk=pk)

        body = (request.POST.get("body") or "").strip()
        if body and len(body) <= 5000:
            SupportMessage.objects.create(ticket=ticket, author=request.user, body=body)
            messages.success(request, "Ответ отправлен.")
        elif body:
            messages.error(request, "Сообщение слишком длинное.")
        return redirect("dashboard:support_detail", pk=pk)

